"""
TRIGO Box Lookup — FastAPI Web Server
--------------------------------------
Local :
    python server.py

Cloud (Render) — variables d'environnement requises :
    DATABASE_URL   → PostgreSQL fourni par Render
    CLOUDINARY_URL → cloudinary://key:secret@cloud_name
"""
import io
import json
import os
import shutil
import socket
import threading
import webbrowser
from datetime import datetime
from pathlib import Path

import qrcode as _qrcode
import uvicorn
from fastapi import FastAPI, Query, Body, UploadFile, File, Form
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles

from box_core import load_mapping, extract_reference, normalize

BASE        = Path(__file__).parent
MANUAL_FILE = BASE / "manual_entries.json"
HISTORY_FILE = BASE / "scan_history.json"
XLSX_PATH   = BASE / "Liste emballages ref Trigo.xlsx"
XLSX_SHEET  = "Feuil3"
HISTORY_MAX = 2000
PORT        = int(os.environ.get("PORT", 8000))

# ── Détection du mode cloud ────────────────────────────────────────────────────

_USE_DB         = bool(os.environ.get("DATABASE_URL"))
_USE_CLOUDINARY = bool(os.environ.get("CLOUDINARY_URL"))

if _USE_DB:
    import database as _db
    _db.init()

if _USE_CLOUDINARY:
    import cloudstore as _cs

# ── Helpers JSON locaux (mode local uniquement) ───────────────────────────────

def _read_json(path: Path, default):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return default
    return default

def _write_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

# ── Application ────────────────────────────────────────────────────────────────

app = FastAPI(title="TRIGO Box Lookup")

# ── Chargement données Excel ───────────────────────────────────────────────────

mapping, original_refs, _, emplacement_map = load_mapping()

# Chargement des entrées manuelles dans le mapping en mémoire
def _load_manual_into_mapping(entries: list[dict]) -> None:
    for e in entries:
        nref = normalize(e["ref"])
        mapping[nref] = e["uc"]
        if e["ref"] not in original_refs:
            original_refs.append(e["ref"])
        if e.get("emplacement"):
            emplacement_map[nref] = e["emplacement"]

if _USE_DB:
    _load_manual_into_mapping(_db.manual_list())
else:
    _manual: list[dict] = _read_json(MANUAL_FILE, [])
    _load_manual_into_mapping(_manual)

known_norm = set(mapping.keys())

# Compteur de scans
if _USE_DB:
    _scan_count: int = _db.history_count()
else:
    _history: list[dict] = _read_json(HISTORY_FILE, [])
    _scan_count: int = len(_history)

# ── Fichiers statiques ─────────────────────────────────────────────────────────

_images_dir = BASE / "box_images"
if _images_dir.exists() and not _USE_CLOUDINARY:
    app.mount("/box_images", StaticFiles(directory=_images_dir), name="box_images")

_placement_images_dir = BASE / "placement_images"
if not _USE_CLOUDINARY:
    _placement_images_dir.mkdir(exist_ok=True)
    app.mount("/placement_images", StaticFiles(directory=_placement_images_dir), name="placement_images")

app.mount("/static", StaticFiles(directory=BASE / "static"), name="static")

# ── Sauvegarde Excel (mode local uniquement) ──────────────────────────────────

def _save_entry_to_excel(ref: str, uc: str, emplacement: str = "", photo_filename: str = "") -> bool:
    if _USE_DB:
        return False
    try:
        import openpyxl
        if not XLSX_PATH.exists():
            return False
        wb = openpyxl.load_workbook(XLSX_PATH)
        if XLSX_SHEET not in wb.sheetnames:
            return False
        ws = wb[XLSX_SHEET]
        headers: dict[str, int] = {
            cell.value: cell.column for cell in ws[1] if cell.value is not None
        }
        for col_name in ("Emplacement", "Photo_Emplacement"):
            if col_name not in headers:
                next_idx = (max(headers.values()) if headers else 0) + 1
                ws.cell(row=1, column=next_idx, value=col_name)
                headers[col_name] = next_idx
        ref_col_idx   = headers.get("Reference", 1)
        uc_col_idx    = headers.get("UC", 2)
        empl_col_idx  = headers["Emplacement"]
        photo_col_idx = headers["Photo_Emplacement"]
        nref = normalize(ref)
        found_row = None
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=ref_col_idx).value
            if cell_val and normalize(str(cell_val)) == nref:
                found_row = row_idx
                break
        if found_row:
            if emplacement:
                ws.cell(row=found_row, column=empl_col_idx, value=emplacement)
            if photo_filename:
                ws.cell(row=found_row, column=photo_col_idx, value=photo_filename)
        else:
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=ref_col_idx,   value=ref)
            ws.cell(row=new_row, column=uc_col_idx,    value=uc)
            ws.cell(row=new_row, column=empl_col_idx,  value=emplacement)
            ws.cell(row=new_row, column=photo_col_idx, value=photo_filename)
        wb.save(XLSX_PATH)
        return True
    except Exception as exc:
        print(f"[Excel] Erreur sauvegarde : {exc}")
        return False

# ── Helpers image ──────────────────────────────────────────────────────────────

def _get_box_image_url(uc: str) -> str | None:
    if _USE_CLOUDINARY and _USE_DB:
        return _db.image_get(f"box:{uc}")
    img_path = _images_dir / f"{uc}.jpg"
    return f"/box_images/{uc}.jpg" if img_path.exists() else None

def _get_placement_image_urls(nref: str) -> list[str]:
    """Retourne la liste de toutes les photos d'emplacement pour une référence."""
    if _USE_CLOUDINARY and _USE_DB:
        url = _db.image_get(f"placement:{nref}")
        return [url] if url else []
    photos = sorted(_placement_images_dir.glob(f"{nref}_*.jpg"), key=lambda p: p.stem)
    return [f"/placement_images/{p.name}" for p in photos]

# ── Helpers historique ─────────────────────────────────────────────────────────

def _history_insert(entry: dict) -> None:
    if _USE_DB:
        _db.history_insert(entry)
    else:
        _history.insert(0, entry)
        if len(_history) > HISTORY_MAX:
            del _history[HISTORY_MAX:]
        _write_json(HISTORY_FILE, _history)

def _local_ip() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

# ── Routes API ─────────────────────────────────────────────────────────────────

@app.get("/")
def index():
    return FileResponse(BASE / "static" / "index.html")


@app.get("/api/stats")
def stats():
    return {"ref_count": len(mapping), "scan_count": _scan_count}


@app.get("/api/lookup")
def lookup(ref: str = Query(...)):
    global _scan_count
    ref = ref.strip()
    if not ref:
        return JSONResponse({"status": "error", "message": "Référence vide",
                             "uc": None, "ref": None, "image_url": None,
                             "emplacement": None, "placement_image_urls": []})

    _scan_count += 1
    now = datetime.now()
    nref = extract_reference(ref, known_norm, original_refs)

    if not nref:
        entry = {
            "datetime": now.isoformat(timespec="seconds"),
            "date": now.strftime("%d/%m/%Y"),
            "time": now.strftime("%H:%M:%S"),
            "raw": ref, "ref": None, "uc": None,
            "status": "error", "message": "Référence non reconnue",
            "image_url": None, "emplacement": None, "placement_image_urls": [],
        }
        _history_insert(entry)
        return JSONResponse(entry)

    uc = mapping.get(nref)
    emplacement = emplacement_map.get(nref)

    entry = {
        "datetime": now.isoformat(timespec="seconds"),
        "date": now.strftime("%d/%m/%Y"),
        "time": now.strftime("%H:%M:%S"),
        "raw": ref,
        "ref": nref,
        "uc": uc,
        "status": "found" if uc else "incomplete",
        "message": "Emballage UC identifié" if uc else "UC non défini",
        "image_url":            _get_box_image_url(uc) if uc else None,
        "emplacement":          emplacement,
        "placement_image_urls": _get_placement_image_urls(nref),
    }
    _history_insert(entry)
    return JSONResponse(entry)


@app.get("/api/history")
def history(
    limit: int = Query(100, ge=1, le=HISTORY_MAX),
    status: str = Query(""),
    q: str = Query(""),
    date: str = Query(""),
):
    if _USE_DB:
        return _db.history_get(limit, status, q, date)
    results = _history
    if status:
        results = [e for e in results if e.get("status") == status]
    if date:
        results = [e for e in results if e.get("date") == date]
    if q:
        ql = q.lower()
        results = [e for e in results if
                   ql in (e.get("ref") or "").lower() or
                   ql in (e.get("raw") or "").lower() or
                   ql in (e.get("uc") or "").lower()]
    return results[:limit]


@app.delete("/api/history")
def clear_history():
    global _scan_count
    _scan_count = 0
    if _USE_DB:
        _db.history_clear()
    else:
        _history.clear()
        _write_json(HISTORY_FILE, [])
    return JSONResponse({"status": "cleared"})


@app.get("/api/network-info")
def network_info():
    ip = _local_ip()
    return {"ip": ip, "port": PORT, "url": f"http://{ip}:{PORT}"}


@app.get("/api/qrcode.png")
def qr_image():
    ip  = _local_ip()
    url = f"http://{ip}:{PORT}"
    img = _qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return Response(content=buf.getvalue(), media_type="image/png")


@app.get("/api/manual")
def manual_list_route():
    if _USE_DB:
        return _db.manual_list()
    return _manual


@app.get("/api/refs")
def refs_search(q: str = Query("", alias="q")):
    q_norm = normalize(q) if q.strip() else ""
    results = []
    for nref, uc in mapping.items():
        if not q_norm or q_norm in nref or (uc and q_norm in normalize(uc)):
            results.append({
                "ref": nref,
                "uc": uc or "",
                "emplacement": emplacement_map.get(nref, ""),
            })
    return sorted(results, key=lambda x: x["ref"])[:300]


@app.post("/api/add")
def add_entry(data: dict = Body(...)):
    ref         = str(data.get("ref",         "")).strip()
    uc          = str(data.get("uc",          "")).strip()
    desc        = str(data.get("desc",        "")).strip()
    emplacement = str(data.get("emplacement", "")).strip()

    if not ref or not uc:
        return JSONResponse(
            {"status": "error", "message": "Référence et UC sont obligatoires"},
            status_code=400,
        )

    nref = normalize(ref)
    mapping[nref] = uc
    known_norm.add(nref)
    if ref not in original_refs:
        original_refs.append(ref)
    if emplacement:
        emplacement_map[nref] = emplacement

    if _USE_DB:
        action = _db.manual_upsert(ref, uc, desc, emplacement)
        return JSONResponse({"status": action, "ref": nref, "uc": uc})

    for existing in _manual:
        if normalize(existing["ref"]) == nref:
            existing.update({"uc": uc, "desc": desc, "emplacement": emplacement})
            _write_json(MANUAL_FILE, _manual)
            _save_entry_to_excel(ref, uc, emplacement)
            return JSONResponse({"status": "updated", "ref": nref, "uc": uc})

    _manual.append({"ref": ref, "uc": uc, "desc": desc, "emplacement": emplacement})
    _write_json(MANUAL_FILE, _manual)
    _save_entry_to_excel(ref, uc, emplacement)
    return JSONResponse({"status": "added", "ref": nref, "uc": uc})


@app.delete("/api/manual/{ref}")
def delete_entry(ref: str):
    nref = normalize(ref)

    if _USE_DB:
        if not _db.manual_delete(ref):
            return JSONResponse({"status": "error", "message": "Entrée introuvable"}, status_code=404)
    else:
        global _manual
        before = len(_manual)
        _manual = [e for e in _manual if normalize(e["ref"]) != nref]
        if len(_manual) == before:
            return JSONResponse({"status": "error", "message": "Entrée introuvable"}, status_code=404)
        _write_json(MANUAL_FILE, _manual)

    mapping.pop(nref, None)
    known_norm.discard(nref)
    emplacement_map.pop(nref, None)
    return JSONResponse({"status": "deleted"})


@app.post("/api/upload-image")
async def upload_image(uc: str = Form(...), file: UploadFile = File(...)):
    if not file.content_type.startswith("image/"):
        return JSONResponse({"status": "error", "message": "Fichier image requis"}, status_code=400)

    file_bytes = await file.read()

    if _USE_CLOUDINARY and _USE_DB:
        url = _cs.upload_box(uc, file_bytes)
        _db.image_set(f"box:{uc}", url, "box")
        return JSONResponse({"status": "ok", "image_url": url})

    _images_dir.mkdir(exist_ok=True)
    dest = _images_dir / f"{uc}.jpg"
    dest.write_bytes(file_bytes)
    return JSONResponse({"status": "ok", "image_url": f"/box_images/{uc}.jpg"})


@app.post("/api/upload-placement")
async def upload_placement(ref: str = Form(...), file: UploadFile = File(...)):
    """Ajoute une photo d'emplacement (indexée) pour une référence."""
    if not file.content_type.startswith("image/"):
        return JSONResponse({"status": "error", "message": "Fichier image requis"}, status_code=400)

    nref = normalize(ref)
    file_bytes = await file.read()

    if _USE_CLOUDINARY and _USE_DB:
        url = _cs.upload_placement(nref, file_bytes)
        _db.image_set(f"placement:{nref}", url, "placement")
        return JSONResponse({"status": "ok", "placement_image_url": url})

    # Trouver le prochain index disponible
    existing = sorted(_placement_images_dir.glob(f"{nref}_*.jpg"))
    next_idx  = len(existing) + 1
    filename  = f"{nref}_{next_idx}.jpg"
    dest      = _placement_images_dir / filename
    dest.write_bytes(file_bytes)

    # Mettre à jour Excel avec tous les fichiers actuels (liste séparée par virgule)
    all_files = ",".join(p.name for p in sorted(_placement_images_dir.glob(f"{nref}_*.jpg")))
    _save_entry_to_excel(ref, mapping.get(nref, ""), emplacement_map.get(nref, ""), all_files)

    return JSONResponse({
        "status": "ok",
        "placement_image_url":  f"/placement_images/{filename}",
        "placement_image_urls": _get_placement_image_urls(nref),
    })


@app.get("/api/placement-photos/{ref}")
def placement_photos(ref: str):
    """Liste toutes les photos d'emplacement d'une référence."""
    nref = normalize(ref)
    return _get_placement_image_urls(nref)


@app.delete("/api/delete-placement/{ref}/{filename}")
def delete_placement_photo(ref: str, filename: str):
    """Supprime une photo d'emplacement spécifique."""
    nref = normalize(ref)
    # Sécurité : le fichier doit appartenir à cette référence
    if not filename.startswith(nref + "_") or ".." in filename or "/" in filename:
        return JSONResponse({"status": "error", "message": "Fichier invalide"}, status_code=400)
    path = _placement_images_dir / filename
    if path.exists():
        path.unlink()
    all_files = ",".join(p.name for p in sorted(_placement_images_dir.glob(f"{nref}_*.jpg")))
    _save_entry_to_excel(ref, mapping.get(nref, ""), emplacement_map.get(nref, ""), all_files)
    return JSONResponse({"status": "deleted", "placement_image_urls": _get_placement_image_urls(nref)})


# ── Lancement local ────────────────────────────────────────────────────────────

def _open_browser():
    import time
    time.sleep(0.9)
    webbrowser.open(f"http://localhost:{PORT}")


if __name__ == "__main__":
    ip = _local_ip()
    print(f"\n  PC     : http://localhost:{PORT}")
    print(f"  Réseau : http://{ip}:{PORT}\n")
    if not _USE_DB:
        threading.Thread(target=_open_browser, daemon=True).start()
    uvicorn.run(app, host="0.0.0.0", port=PORT)
